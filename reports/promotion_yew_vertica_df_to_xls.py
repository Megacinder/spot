from dataclasses import dataclass
from pandas import DataFrame
from typing import Tuple

from secret.sql.reports_for_yew import (
    Vietnam
)

from secret.sql.reports_for_yew_Jan23 import (
    SQL_USERS, SQL_PARTNERS, SQL_GALA_INDIA, SQL_GALA_TH, SQL_GALA_TH_ALL_CLIENTS,
    ThLaMa, ThLaMaGala, IndiaGala, India, ThLaMaGalaAllClients,
)

from src.connection.envs import envs
from src.connection.vertica import VerticaCursor

envs = envs()


def take_params(data_class: dataclass, template_sql: str, type: tuple = ('user',)) -> Tuple[str, str]:
    dc = data_class()
    if 'jan23' in type:
        country_list = "'" + "','".join(str(i) for i in dc.country_list) + "'"
        sql = template_sql.format(
            from_dt=dc.from_dt,
            to_dt=dc.to_dt,
            country_list=country_list,
        )
        filename = dc.filename
    else:
        partner_user_id = ','.join(str(i) for i in dc.partner_user_id)
        sql = template_sql.format(
            partner_user_id=partner_user_id,
            from_dt=dc.from_dt,
            to_dt=dc.to_dt,
        )
        dict1 = {
            'user': dc.filename,
            'partner': dc.filename_partner,
        }
        filename = dict1[type[0]]

    return sql, filename


def execute_sql(cur, sql: str) -> DataFrame:
    cur.execute(sql)

    df = DataFrame(cur.fetchall())
    cols = [col[0] for col in cur.description]

    if not df.empty:
        df.columns = cols

    for col in cols:
        if 'usd' in col.lower() or 'volume' in col.lower():
            df[col] = df[col].astype(float)
            df[col] = df[col].round(2)

    return df


def auto_adjust_column_widths(excel_file: "Excel File Path", extra_space=1) -> None:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    wb = load_workbook(excel_file)

    for ws in wb:
        df = DataFrame(ws.values, )

        for i, r in (df.astype(str).applymap(len).max(axis=0) + extra_space).items():
            ws.column_dimensions[get_column_letter(i + 1)].width = r

    wb.save(excel_file)


with VerticaCursor() as cur:

    classes = (ThLaMa, India, Vietnam)
    for dc in classes:
        sql, filename = take_params(dc, SQL_USERS, ('user',))
        df = execute_sql(cur, sql)
        df.to_excel(envs['PATH_FOR_FILES'] + filename)
        auto_adjust_column_widths(envs['PATH_FOR_FILES'] + filename)

        sql, filename = take_params(dc, SQL_PARTNERS, ('partner',))
        df = execute_sql(cur, sql)
        df.to_excel(envs['PATH_FOR_FILES'] + filename)
        auto_adjust_column_widths(envs['PATH_FOR_FILES'] + filename)

    classes_and_scripts = {
        ThLaMaGala: SQL_GALA_TH,
        IndiaGala: SQL_GALA_INDIA,
        ThLaMaGalaAllClients: SQL_GALA_TH_ALL_CLIENTS,
    }
    for dc, sql_script in classes_and_scripts.items():
        sql, filename = take_params(dc, sql_script, ('jan23', 'gala',))
        df = execute_sql(cur, sql)
        df.to_excel(envs['PATH_FOR_FILES'] + filename)
        auto_adjust_column_widths(envs['PATH_FOR_FILES'] + filename)
