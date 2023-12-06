from pathlib import Path
from pandas import DataFrame
from typing import Dict, Any, Union, Mapping

def get_filename(path: str) -> str:
    return Path(path).name.replace('.py', '')


def auto_adjust_column_widths(excel_file: "Excel File Path", extra_space: int = 1) -> None:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter

    wb = load_workbook(excel_file)

    for ws in wb:
        df1 = DataFrame(ws.values, )
        for i, r in (df1.astype(str).applymap(len).max(axis=0) + extra_space).items():
            ws.column_dimensions[get_column_letter(i + 1)].width = r

    wb.save(excel_file)


def update(source: Dict[Any, Any], overrides: Union[Dict[Any, Any], Mapping]):
    for key, val in overrides.items():
        if isinstance(val, Mapping):
            tmp = update(source.get(key, {}), val)
            source[key] = tmp
        elif isinstance(val, list):
            source[key] = (source.get(key, []) + val)
        else:
            source[key] = overrides[key]
    return source


def get_airflow_dag_id(path: str) -> str:
    return Path(path).name.replace('.py', '')


def get_current_path(path: str) -> str:
    return Path(path).parent.name
